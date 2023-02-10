import time, json, socket
from openpyxl import load_workbook
from ast import literal_eval
import matplotlib.pyplot as plt

SIZE = 2048
ENCODING = 'utf-8'

def sendRequestV2(s, tag, request, delay=0):
    """ send request to server
    Args:
        s (socket): client's socket
        tag (str): request tag
        request (dict): request parameters
    Returns:
    """
    # add tag to request
    request['request'] = tag

    # send request
    time.sleep(delay)
    s.sendall(bytes(json.dumps(request) + "\r\n", ENCODING))
    # print(f"[{tag}] Send request (no response)")

def sendRequestAndReceiveV2(s, tag, request, delay=0):
    """ send request and receive response to/from server
    Args:
        s (socket): client's socket
        tag (str): request tag
        request (dict): request parameters
    Returns:
        dict: response (json)
    """
    # add tag to request
    request['request'] = tag

    # send request
    time.sleep(delay)
    s.sendall(bytes(json.dumps(request) + "\r\n", ENCODING))
    # print(f"[{tag}] Send request")
    # print(f"[{tag}] Send {request}")

    # receive server response
    # response must ends with "\r\n"
    receivedStr = ''
    flag = False
    while True:
        try:
            received = str(s.recv(SIZE), ENCODING)
            flag = True
            if received.endswith("\r\n"):
                received = received[:-2]
                receivedStr = receivedStr + received
                break
            receivedStr = receivedStr + received
        except socket.timeout:
            if flag:
                if receivedStr.endswith("\r\n"):
                    receivedStr = receivedStr[:-2]
                break

    try:
        response = json.loads(receivedStr)
        # print(f"[{tag}] Receive response")
        # print(f"[{tag}] receive response {response}")
        return response
    except json.decoder.JSONDecodeError:
        #raise Exception(f"[{tag}] Server response with: {response}")
        print(f"[{tag}] Server response with: {receivedStr}")

    # Do not close socket to maintain the connection.

# deprecated
def listenAndAccept(mq, serverSocket, SIZE, ENCODING):
    """ listen and accept from server socket and put in message queue
    Args:
        mq (MessageQueue): message queue
        serverSocket (socket): server's socker
        SIZE (int): buffer size
        ENCODING (str): encoding
    Returns:
    """
    serverSocket.listen(200)

    while True:  # always listen
        clientSocket, addr = serverSocket.accept()

        # receive client data
        # client request must ends with "\r\n"
        request = ''
        while True:
            received = str(clientSocket.recv(SIZE), ENCODING)
            if received.endswith("\r\n"):
                received = received.replace("\r\n", "")
                request = request + received
                break
            request = request + received

        mq.put(clientSocket, request)

def writeToExcel(filename, run_data):
    write_wb = load_workbook(filename)
    write_ws = write_wb.create_sheet(str(int(time.time())))
    for data in run_data:
        write_ws.append(data)
    write_wb.save(filename)
    write_wb.close()

def writeWeightsToFile(weights):
    # WARN: STATIC FILE PATH!
    # weights must be 1-dim list
    f = open('../../results/model.txt', 'w')
    f.write(str(weights))
    f.close()

def readWeightsFromFile():
    # WARN: STATIC FILE PATH!
    # return 1-dim list (weights)
    f = open('../../results/model.txt', 'r')
    weights = f.readline()
    f.close()
    return literal_eval(weights)

def plotResults():
    filename = '../results/최종/24-27.xlsx'
    write_wb = load_workbook(filename, read_only=True, data_only=True)
    r = [x for x in range(100)]
    acc = [
        [x[0].value for x in write_wb['24']['B1':'B100']],
        [x[0].value for x in write_wb['25']['B1':'B100']],
        [x[0].value for x in write_wb['4']['B1':'B100']],
        [x[0].value for x in write_wb['26']['B1':'B100']],
        [x[0].value for x in write_wb['27']['B1':'B100']],
        [x[0].value for x in write_wb['10']['B1':'B100']],
    ]

    plt.plot(r, acc[0], label='C=1, qLevel=30')
    plt.plot(r, acc[1], label='C=1, qLevel=100')
    plt.plot(r, acc[2], label='C=1, qLevel=300')
    plt.plot(r, acc[3], label='C=4, qLevel=30')
    plt.plot(r, acc[4], label='C=4, qLevel=100')
    plt.plot(r, acc[5], label='C=4, qLevel=300')
    plt.legend()
    plt.xlabel('Round')
    plt.ylabel('Test Accuracy(%)')
    plt.show()

    write_wb.close()


if __name__ == "__main__":
    plotResults()
